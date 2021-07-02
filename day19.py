import openpyxl
path = "studdata.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
cell_obj = sheet_obj.cell(row = 5, column = 3)
print(cell_obj.value)


for i in range(1,12):
    cell_obj = sheet_obj.cell(row = 5, column = i)
    print(cell_obj.value)

    import mysql.connector

mydb = mysql.connector.connect(
  host="localhost",
 user="root",
  password="jes29",
)

mycursor = mydb.cursor()
print(mydb)

dbse = mydb.cursor()

dbse.execute("CREATE DATABASE Students_management_database")
dbse = mydb.cursor()

dbse.execute("SHOW DATABASES")

for entry in dbse:
  print(entry)

  mydb = mysql.connector.connect(
  host="localhost",
 user="root",
  password="0712",
  database="SMD"
)
dbse = mydb.cursor()

dbse.execute("CREATE TABLE studentdata (rollno INT(8),Age INT(2),NAME VARCHAR(255), sub1 INT(25),sub2 INT(25),sub3 INT(25),sub4 INT(3), total INT(20) ,email_id VARCHAR(255))")

dbse = mydb.cursor()

dbse.execute("SHOW TABLES")

for value in dbse:
  print(value)

  cur = mydb.cursor()
cur.execute('SELECT * FROM studentdata')
for row in cur:
    print(row)

    import pandas as pd

df = pd.read_excel('studdata.xlsx')

import xlrd
import MySQLdb

xl_sheet = xlrd.open_workbook("studdata.xlsx")
xl_sheet

sheet_name =xl_sheet.sheet_names()
sheet_name

mydb = mysql.connector.connect(
  host="localhost",
 user="root",
  password="jes29",
  database="students_management_database"
)

cur = mydb.cursor()
for x in range(0,1):
    sheet=xl_sheet.sheet_by_index(x)
    sql= "INSERT INTO studentdata(roll_no,Age,NAME,sub1,sub2 ,sub3 ,sub4, total,email_id) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s)"
    for r in range(1,sheet.nrows):
        roll_no =sheet.cell(r,0).value
        Age =sheet.cell(r,1).value
        NAME =sheet.cell(r,2).value
        sub1 =sheet.cell(r,3).value
        sub2 =sheet.cell(r,4).value
        sub3 =sheet.cell(r,5).value
	      sub4 =sheet.cell(r,6).value
        total =sheet.cell(r,7).value
        email_id=sheet.cell(r,8).value
        values =(roll_no ,Age,NAME ,sub1,sub2 ,sub3 ,sub4, total ,email_id)
        
        cur.execute(sql,values)
mydb.commit()

mycursor = mydb.cursor()

mycursor.execute("SELECT * FROM studentdata")

myresult = mycursor.fetchall()

for i in myresult:
    print(i)

    mycursor = mydb.cursor()

mycursor.execute("SELECT NAME FROM studentdata WHERE total >300")

myresult = mycursor.fetchall()

for i in myresult:
    print(i)

    mydb.commit()

mydb.close()

